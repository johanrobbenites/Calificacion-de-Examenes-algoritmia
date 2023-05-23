import subprocess
from openpyxl import Workbook

archivo_cpp = "CALDERON-ANTHONY-P1.cpp"
nombre_reporte = "reporte_cpplint.xlsx"

# Ejecutar el comando de cpplint y capturar la salida
comando = ["cpplint", archivo_cpp]
resultado = subprocess.run(comando, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

# Almacenar los errores de cpplint en una lista
errores = resultado.stderr.splitlines()

# Crear un libro de Excel y una hoja para almacenar los resultados
libro = Workbook()
hoja = libro.active
hoja.title = "Reporte de cpplint"

# Escribir los encabezados de la tabla en la hoja
hoja.cell(row=1, column=1, value="Archivo")
hoja.cell(row=1, column=2, value="Errores")

# Escribir el nombre del archivo y el número de errores en la hoja
hoja.cell(row=2, column=1, value=archivo_cpp)
hoja.cell(row=2, column=2, value=len(errores))

# Guardar el libro en un archivo
libro.save(nombre_reporte)

print("Reporte generado exitosamente en " + nombre_reporte)