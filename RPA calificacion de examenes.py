import subprocess
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

#Paso 1: Ingreso de las rutas
ruta_nombres = r".\EXAMEN -PC-CALIFICADAS"
ruta_EF = r".\Final"
ruta_PC4 = r".\PC4"
ruta_Sustitutorio = r".\Sustitutorio"
soluciones = r".\Soluciones"

#Paso 2: Obtenciond e apellidos y nombres de alumnos 
nombre_alumnos = [nombre for nombre in os.listdir(ruta_nombres) if os.path.isdir(os.path.join(ruta_nombres, nombre))] 
apellidos = [nombre.split(" ")[0] for nombre in os.listdir(ruta_nombres) if os.path.isdir(os.path.join(ruta_nombres, nombre))]
nombres = [nombre.split()[2] for nombre in nombre_alumnos if len(nombre.split()) >= 3]


#Paso 3: Creacion de funciones de leer archivo c++
def contar_lineas_cpp(rutas):
    num_lineas = []
    for ruta in rutas:
        with open(ruta, 'r') as archivo:
            lineas = archivo.readlines()
            num_lineas.append(len(lineas))
    return num_lineas

archivos_EF =  os.listdir(ruta_EF)
archivos_PC4 =  os.listdir(ruta_PC4)
archivos_Sustitutorios =  os.listdir(ruta_Sustitutorio)

def calificacion(numero):
    rango_valido = 350
    posicion = (numero + 100) / rango_valido
    calificacion = round(posicion * 20)
    return calificacion


notas_EF =[]
obs_EF=[]
cods_EF=[]
lineas_EF=[]

notas_PC4=[]
obs_PC4=[]
cods_PC4=[]
lineas_PC4=[]

notas_Sustitutorio=[]
cods_Sust=[]
obs_Sust=[]
lineas_Sust=[]


nombre_reporte = "Reporte.xlsx"

libro = Workbook()
hoja = libro.active
hoja.title = "Acta de Notas"

# Paso 4: Escribir los encabezados de la tabla en la hoja
hoja.cell(row=1, column=1, value="Alumno").font = Font(bold=True)
hoja.cell(row=1, column=2, value="PC4").font = Font(bold=True)
hoja.cell(row=1, column=3, value="EF").font = Font(bold=True)
hoja.cell(row=1, column=4, value="ES").font = Font(bold=True)
hoja.cell(row=1, column=5, value="Nota Final").font = Font(bold=True)


for i, alumno in enumerate(nombre_alumnos):
    hoja.cell(row=i+2, column=1, value=alumno)
    hoja.cell(row=i+2, column=2, value="NSP").alignment = Alignment(horizontal='center')
    obs_PC4.append("No hay codigo")
    notas_PC4.append(0)
    hoja.cell(row=i+2, column=3, value="NSP").alignment = Alignment(horizontal='center')
    notas_EF.append(0)
    obs_EF.append("No hay codigo")
    hoja.cell(row=i+2, column=4, value="NSP").alignment = Alignment(horizontal='center')
    obs_Sust.append("No hay codigo")
    notas_Sustitutorio.append(0)

# Paso 5: Creacion de funciones para la revision de la PC4, EF y ES
for i, archivo_EF in enumerate(archivos_EF):
    path_archivo = ruta_EF + "\\" + archivo_EF
    cods_EF.append(path_archivo)
    sol = [soluciones]
    comando = ["cpplint", path_archivo]
    resultado = subprocess.run(comando, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    errores = resultado.stderr.splitlines()
    apellido = archivo_EF.split('-')[0].upper()
    nombre = archivo_EF.split('-')[1].upper()
    contador=0
    if(apellido in apellidos and nombre in nombres):
        for j, apellidover in enumerate(apellidos):
            if(apellidover==apellido):
                contador=j
                break
        hoja.cell(row=contador+2, column=3, value=len(errores)).alignment = Alignment(horizontal='center')
        obs_EF[contador]=errores[-3].split()[::]
        notas_EF[contador]=len(errores)
    else:
        continue
num_lineas = contar_lineas_cpp(cods_EF)
lineas_EF=num_lineas
num_lineas=[]

iteradorEF=0
for i, notasve in enumerate(notas_EF):
    if(notasve!=0):

        hoja.cell(row=i+2, column=3, value=calificacion(lineas_EF[iteradorEF]*3-notasve)).alignment = Alignment(horizontal='center')
        iteradorEF+=1

for i in range(len(obs_EF)):
    obs_EF[i] = str(obs_EF[i])

for i, archivo_PC4 in enumerate(archivos_PC4):
    path_archivo = ruta_PC4 + "\\" + archivo_PC4
    cods_PC4.append(path_archivo)
    comando = ["cpplint", path_archivo]
    resultado = subprocess.run(comando, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    errores = resultado.stderr.splitlines()
    apellido = archivo_PC4.split('-')[0].upper()
    nombre = archivo_PC4.split('-')[1].upper()
    contador=0
    if(apellido in apellidos and nombre in nombres):
        for j, apellidover in enumerate(apellidos):
            if(apellidover==apellido):
                contador=j
                break
        hoja.cell(row=contador+2, column=2, value=len(errores))
        obs_PC4[contador]=errores[-3].split()[2:]
        notas_PC4[contador]=len(errores)    
    else:
        continue

num_lineas = contar_lineas_cpp(cods_PC4)
lineas_PC4=num_lineas
num_lineas=[]

iteradorPC4=0
for i, notasve in enumerate(notas_PC4):
    if(notasve!=0):
        hoja.cell(row=i+2, column=2, value=calificacion(lineas_PC4[iteradorPC4]*3-notasve)).alignment = Alignment(horizontal='center')
        iteradorPC4+=1


for i in range(len(obs_PC4)):
    obs_PC4[i] = str(obs_PC4[i])

for i, archivo_sustitutorio in enumerate(archivos_Sustitutorios):
    path_archivo = ruta_Sustitutorio + "\\" + archivo_sustitutorio
    cods_Sust.append(path_archivo)
    comando = ["cpplint", path_archivo]
    resultado = subprocess.run(comando, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

    errores = resultado.stderr.splitlines()
    apellido = archivo_sustitutorio.split('-')[0].upper()
    nombre = archivo_sustitutorio.split('-')[1].upper()
    contador=0
    if(apellido in apellidos and nombre in nombres):
        for j, apellidover in enumerate(apellidos):
            if(apellidover==apellido):
                contador=j
                break
        hoja.cell(row=contador+2, column=4, value=len(errores)).alignment = Alignment(horizontal='center')
        obs_Sust[contador]=errores[-3].split()[2:]
        notas_Sustitutorio[contador]=len(errores)
    else:
            continue
num_lineas = contar_lineas_cpp(cods_Sust)
lineas_Sust=num_lineas
num_lineas=[]

iteradorSust=0
for i, notasve in enumerate(notas_Sustitutorio):
    if(notasve!=0):
        hoja.cell(row=i+2, column=4, value=calificacion(lineas_Sust[iteradorSust]*3-notasve)).alignment = Alignment(horizontal='center')
        iteradorSust+=1

for i in range(len(obs_Sust)):
    obs_Sust[i] = str(obs_Sust[i])
c1=0
c2=0
c3=0
notas_finales=[]
#calculo de las notas finales
for i in range(len(notas_EF)):
    ef=0
    es=0
    pc4n=0
    nf=0
    if(notas_EF[i]!=0):
        ef=calificacion(lineas_EF[c1]*3-notas_EF[i])
        c1+=1
    if(notas_Sustitutorio[i]!=0):
        es=calificacion(lineas_Sust[c2]*3-notas_Sustitutorio[i])
        c2+=1
    if(notas_PC4[i]!=0):
        pc4n=calificacion(lineas_PC4[c3]*3-notas_PC4[i])
        c3+=1
    if(es<ef and es!=0):
        nf=(es*2+pc4n)/3
    else:
        nf=(ef*2+pc4n)/3
    notas_finales.append(nf)

for i,nota_final in enumerate(notas_finales):
    hoja.cell(row=i+2, column=5, value=nota_final).alignment = Alignment(horizontal='center')

hoja2 = libro.create_sheet(title="Posibles Plagios")
hoja2.cell(row=1, column=1, value="Alumno").font = Font(bold=True)
hoja2.cell(row=1, column=2, value="Plagio PC4").font = Font(bold=True)
hoja2.cell(row=1, column=3, value="Plagio EF").font = Font(bold=True)
hoja2.cell(row=1, column=4, value="Plagio sustitutorio").font = Font(bold=True)


# Paso 6: Verificacion de plagios de las evaluaciones en  c++
for i, alumno in enumerate(nombre_alumnos):
    hoja2.cell(row=i+2, column=1, value=alumno)
    hoja2.cell(row=i+2, column=2, value=" ").alignment = Alignment(horizontal='center')
    hoja2.cell(row=i+2, column=3, value=" ").alignment = Alignment(horizontal='center')
    hoja2.cell(row=i+2, column=4, value=" ").alignment = Alignment(horizontal='center')

plagios_PC4 = list(filter(lambda x: notas_PC4.count(x) > 1, set(notas_PC4)))

plagios_PC4.remove(0)
cont_plag_pc4=1
for plagiado in plagios_PC4:
    for i,notas in enumerate(notas_PC4):
        if(plagiado==notas):
            valor="Plagio "+str(cont_plag_pc4)
            hoja2.cell(row=i+2, column=2, value=valor).alignment = Alignment(horizontal='center')
    cont_plag_pc4+=1

plagios_EF = list(filter(lambda x: notas_EF.count(x) > 1, set(notas_EF)))

plagios_EF.remove(0)

cont_plag_EF=1
plagios_EF.remove(35)
for plagiado in plagios_EF:
    for i,notas in enumerate(notas_EF):
        if(plagiado==notas):
            valor="Plagio "+str(cont_plag_EF)
            hoja2.cell(row=i+2, column=3, value=valor).alignment = Alignment(horizontal='center')
    cont_plag_EF+=1

# Paso 7: Observaciones de los examenes

hoja3 = libro.create_sheet(title="Observaciones")

hoja3.cell(row=1, column=1, value="Alumno").font = Font(bold=True)
hoja3.cell(row=1, column=2, value="# errores PC4").font = Font(bold=True)
hoja3.cell(row=1, column=3, value="Observacion").font = Font(bold=True)
hoja3.cell(row=1, column=4, value="# errores EF").font = Font(bold=True)
hoja3.cell(row=1, column=5, value="Observacion").font = Font(bold=True)
hoja3.cell(row=1, column=6, value="# errores ES").font = Font(bold=True)
hoja3.cell(row=1, column=7, value="Observacion").font = Font(bold=True)

for i, alumno in enumerate(nombre_alumnos):
    hoja3.cell(row=i+2, column=1, value=alumno)
    hoja3.cell(row=i+2, column=2, value=" ").alignment = Alignment(horizontal='center')
    hoja3.cell(row=i+2, column=3, value=" ").alignment = Alignment(horizontal='center')
    hoja3.cell(row=i+2, column=4, value=" ").alignment = Alignment(horizontal='center')
    hoja3.cell(row=i+2, column=5, value=" ").alignment = Alignment(horizontal='center')
    hoja3.cell(row=i+2, column=6, value=" ").alignment = Alignment(horizontal='center')
    hoja3.cell(row=i+2, column=7, value=" ").alignment = Alignment(horizontal='center')

for i,num_pc4 in enumerate(notas_PC4):
    hoja3.cell(row=i+2, column=2, value=num_pc4).alignment = Alignment(horizontal='center')
for i,obs in enumerate(obs_PC4):
    hoja3.cell(row=i+2, column=3, value=obs).alignment = Alignment(horizontal='left')

for i,num_EF in enumerate(notas_EF):
    hoja3.cell(row=i+2, column=4, value=num_EF).alignment = Alignment(horizontal='center')
for i,obs in enumerate(obs_EF):
    hoja3.cell(row=i+2, column=5, value=obs).alignment = Alignment(horizontal='left')

for i,num_susti in enumerate(notas_Sustitutorio):
    hoja3.cell(row=i+2, column=6, value=num_susti).alignment = Alignment(horizontal='center')
for i,obs in enumerate(obs_Sust):
    hoja3.cell(row=i+2, column=7, value=obs).alignment = Alignment(horizontal='left') 


# Paso 8: Poner bordes al codigo Excel
border_style = Side(border_style='thin', color='000000')
for col in hoja.columns:
    for cell in col:
        cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
for col in hoja2.columns:
    for cell in col:
        cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
for col in hoja.columns:
    for cell in col:
        cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

# Paso 9: Guardar e imprimir el reporte
libro.save(nombre_reporte)